
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font,PatternFill,colors
import math
def format_excel_file(path):
    try:
        wb=load_workbook(filename=path)
        ws=wb.active


        #fill cell color of brown to all 1st row cells
        def change_color_row_and_coloumn(position,color):
            for i in ws[str(position)]:
                i.fill=PatternFill(fgColor=color,fill_type="solid")
        def get_maximin_width(column_name):
            max_width=0
            for i in ws[column_name]:
                if len(str(i.value).strip())>max_width:
                    max_width=len(str(i.value))
            return max_width

        ws[f"{get_column_letter(ws.max_column+1)}1"]="Graham Value"
        #Bold the header and capitalize the first letter of each word
        ws[f"{get_column_letter(ws.max_column+1)}1"]="Diff(LTP-Graham Value)"

        for each_row in ws['1']:
            if each_row.value.lower()=="eps":
                eps_position_letter=each_row.column_letter
            if each_row.value.lower()=="book value":
                book_value_position_letter=each_row.column_letter
            if each_row.value.lower()=="market price":
                market_price_position_letter=each_row.column_letter
            if each_row.value.lower()=="graham value".lower():
                graham_value_position_letter=each_row.column_letter
            if each_row.value.lower()=="diff(ltp-graham value)".lower():
                diff_ltp_graham_value_position_letter=each_row.column_letter

        accessing_columns=[graham_value_position_letter,diff_ltp_graham_value_position_letter]
        for accessin_column in accessing_columns:
            row=2
            for each_cell in ws[accessin_column][1:]:
            # print(each_cell)
            # eps=str(eps_position_letter)+str(row)
            # # print(eps)
            # book_value=str(book_value_position_letter)+str(row)
                if ws[f"{accessin_column}1"].value.lower()=="Graham Value".lower():
                    try:
                        graham=math.sqrt(22.5*float(ws[f"{eps_position_letter}{row}"].value)*float(ws[f"{book_value_position_letter}{row}"].value))
                    except:
                        graham=0
                    ws[f"{each_cell.coordinate}"]=graham
                elif ws[f"{accessin_column}1"].value.lower()=="Diff(LTP-Graham Value)".lower():
                    try:
                        diff=float(str(ws[f"{market_price_position_letter}{row}"].value).replace(",",''))-float(ws[f"{graham_value_position_letter}{row}"].value)
                    except:
                        diff=1000
                    ws[f"{each_cell.coordinate}"]=diff
                
                    if(float( ws[f"{each_cell.coordinate}"].value<1)):
                        # print("less than 0")
                        change_color_row_and_coloumn(row,"228B22")
                    elif(float(ws[f"{each_cell.coordinate}"].value<20)):
                        change_color_row_and_coloumn(row,"32CD32")
                    elif(float(ws[f"{each_cell.coordinate}"].value<50)):
                        change_color_row_and_coloumn(row,"00FF00")
                    elif(float(ws[f"{each_cell.coordinate}"].value<100)):
                        change_color_row_and_coloumn(row,"7CFC00")
                row+=1

        change_color_row_and_coloumn('1','FF0000')
        #auto adjust column width to fit content
        for adjust_each_column in ws["1"]:
            ws.column_dimensions[adjust_each_column.column_letter].width=get_maximin_width(adjust_each_column.column_letter)
            
        ws.auto_filter.ref=f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"#configure filter form A1 to last row and last column
        wb.save(path)
        return(True)
    except:
        return(False)
